from pathlib import Path
import pandas as pd
import sys
from openpyxl import load_workbook

EXPECTED_SHEETS = [
    "Dashboard",
    "Summary",
    "Traceability Gaps",
    "Execution_Detail",
    "Missing",
    "Extra",
]

RELEASE = "2026.04"

BASELINE = Path(f"tests/regression/baseline/Traceability_Reconciliation_{RELEASE}.xlsx")
OUTPUT = Path(f"outputs/{RELEASE}/Traceability_Reconciliation_{RELEASE}.xlsx")


def load_sheet(path, sheet):
    return pd.read_excel(path, sheet_name=sheet).fillna("")

def compare_workbook_structure():

    issues = []

    wb = load_workbook(OUTPUT)

    # -------------------------------
    # Sheet existence
    # -------------------------------
    for sheet in EXPECTED_SHEETS:

        if sheet not in wb.sheetnames:
            issues.append(f"Missing sheet: {sheet}")

    # -------------------------------
    # Summary checks
    # -------------------------------
    if "Summary" in wb.sheetnames:

        ws = wb["Summary"]

        if ws.freeze_panes != "A2":
            issues.append("Summary freeze panes missing")

        if not ws.auto_filter.ref:
            issues.append("Summary autofilter missing")

        hyperlink_count = 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    hyperlink_count += 1

        if hyperlink_count == 0:
            issues.append("Summary hyperlinks missing")

    return issues

def compare_sheets(sheet):
    print(f"Comparing sheet: {sheet}")

    df_base = load_sheet(BASELINE, sheet)
    df_new = load_sheet(OUTPUT, sheet)

    common_cols = sorted(set(df_base.columns) & set(df_new.columns))
    df_base = df_base[common_cols]
    df_new = df_new[common_cols]

    df_base = df_base.sort_values(common_cols).reset_index(drop=True)
    df_new = df_new.sort_values(common_cols).reset_index(drop=True)

    # ----------------------------------------------------------
    # Compare sheets
    # ----------------------------------------------------------
    print(f"\n--- {sheet} ---")

    if not df_base.equals(df_new):
        diff = df_base.compare(df_new)

        print(f"❌ {sheet} mismatch")
        print(f"Rows changed: {len(diff)}")

        print("\nFirst differences:")
        print(diff.head(20))

        output_dir = Path("outputs/diff")
        output_dir.mkdir(parents=True, exist_ok=True)

        safe_sheet = sheet.replace(" ", "_")
        diff_file = output_dir / f"diff_{safe_sheet}.xlsx"

        with pd.ExcelWriter(diff_file, engine="openpyxl") as writer:
            diff.to_excel(writer, sheet_name="Diff")

        print(f"\nDiff written to: {diff_file}")

        return False   # 👈 IMPORTANT

    print(f"✅ {sheet} matches")
    return True        # 👈 IMPORTANT

def main():
    print("Running regression...")

    if not OUTPUT.exists():
        print(f"❌ Missing output file: {OUTPUT}")
        sys.exit(1)

    if not BASELINE.exists():
        print(f"❌ Missing baseline file: {BASELINE}")
        sys.exit(1)

    sheets = [
        "Dashboard",
        "Summary",
        "Traceability Gaps",
        "Execution_Detail"
    ]

    results = [compare_sheets(s) for s in sheets]

    data_ok = all(results)

    structure_failures = compare_workbook_structure()

    structure_ok = len(structure_failures) == 0

    # ----------------------------------------------------------
    # Structure failures
    # ----------------------------------------------------------
    if not structure_ok:

        print("\n❌ WORKBOOK STRUCTURE FAILED")

        for issue in structure_failures:
            print(f" - {issue}")

    # ----------------------------------------------------------
    # Final outcome
    # ----------------------------------------------------------
    if not data_ok or not structure_ok:
        sys.exit(1)

    print("\n✅ REGRESSION PASSED")

if __name__ == "__main__":
    main()