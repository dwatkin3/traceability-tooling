from __future__ import annotations
from pathlib import Path
from typing import Iterable
import pandas as pd
import re
from .status_utils import classify_status
from .id_normaliser import normalise_id, normalise_text
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def _format_workbook(workbook):

	for ws in workbook.worksheets:

		# --------------------------------------------------
		# Freeze header row
		# --------------------------------------------------
		ws.freeze_panes = "A2"

		# --------------------------------------------------
		# Enable filters
		# --------------------------------------------------
		if ws.max_row > 1:
			ws.auto_filter.ref = ws.dimensions

		# --------------------------------------------------
		# Auto-size columns
		# --------------------------------------------------
		for column_cells in ws.columns:

			max_length = 0

			for cell in column_cells:

				try:
					value = str(cell.value or "")
				except Exception:
					value = ""

				if len(value) > max_length:
					max_length = len(value)

			column_letter = get_column_letter(
				column_cells[0].column
			)

			# --------------------------------------------------
			# Width rules
			# --------------------------------------------------
			if max_length < 10:
				width = 12
			elif max_length > 80:
				width = 80
			else:
				width = max_length + 2

			ws.column_dimensions[column_letter].width = width

def _add_summary_hyperlinks(wb):

	summary_ws = wb["Summary"]
	gaps_ws = wb["Traceability Gaps"]
	detail_ws = wb["Execution_Detail"]

	gaps_lookup = {}
	detail_lookup = {}

	for row in range(2, gaps_ws.max_row + 1):

		story = gaps_ws[f"B{row}"].value

		if story:
			gaps_lookup[story] = row

	for row in range(2, detail_ws.max_row + 1):

		story = detail_ws[f"B{row}"].value

		if story and story not in detail_lookup:
			detail_lookup[story] = row

	for row in range(2, summary_ws.max_row + 1):

		story = summary_ws[f"B{row}"].value

		if story in gaps_lookup:

			cell = summary_ws[f"C{row}"]

			cell.hyperlink = (
				f"#'Traceability Gaps'!A{gaps_lookup[story]}"
			)

			cell.font = Font(
				color="0000FF",
				underline="single"
			)

		if story in detail_lookup:

			cell = summary_ws[f"D{row}"]

			cell.hyperlink = (
				f"#'Execution_Detail'!A{detail_lookup[story]}"
			)

			cell.font = Font(
				color="0000FF",
				underline="single"
			)

def _df_from_set(name: str, values: Iterable[str]) -> pd.DataFrame:
	return pd.DataFrame({name: sorted(values)})

def _derive_exec_status(
	total_exec,
	passed,
	failed,
	in_progress,
	not_started,
	passed_with_evidence
):
	"""
	Determine STORY-level execution status
	"""

	if total_exec == 0:
		return "🔴 No execution tests"

	if failed > 0:
		return "🔴 Failed tests present"

	if passed > 0 and passed_with_evidence < passed and failed == 0:
		return "🔴 Passed but missing evidence"

	if in_progress > 0:
		return "🟠 In progress"

	if total_exec > 0 and passed == 0 and in_progress == 0:
		return "🟠 Not started"

	if passed == total_exec and passed_with_evidence == total_exec:
		return "🟢 Passed with evidence"

	return "🟠 Mixed / Unknown"

def _build_dashboard(df_summary, df_gaps):

	planned_stories = len(
		df_summary[df_summary["Release"] != "UNREFERENCED"]
	)

	executed_stories = len(
		df_gaps[df_gaps["Coverage Count"] > 0]
	)

	failed_tests = int(df_summary["Failed"].sum())

	missing_tests = int(df_gaps["Missing Count"].sum())

	misaligned_tests = 0

	for val in df_gaps["Misaligned Tests"].fillna(""):

		if val:
			misaligned_tests += len([
				x for x in val.split(",")
				if x.strip()
			])

	planned_total = int(df_gaps["Planned Count"].sum())
	covered_total = int(df_gaps["Coverage Count"].sum())

	coverage_pct = 0

	if planned_total:
		coverage_pct = round(
			(covered_total / planned_total) * 100,
			1
		)

	rows = [
		("Planned stories", planned_stories),
		("Executed stories", executed_stories),
		("Planned Test Coverage %", coverage_pct),
		("Failed tests", failed_tests),
		("Missing tests", missing_tests),
		("Misaligned tests", misaligned_tests),
	]

	return pd.DataFrame(rows, columns=["Metric", "Value"])



def _build_execution_detail(df_exec, release_story_to_tests, story_to_release):
	"""
	Build detailed execution mapping per story + test.

	Assumptions:
	- df_exec is already normalised (IDs cleaned in parser)
	- story/test identifiers should use normalise_id
	- text fields (sheet/file) should use normalise_text
	"""

	rows = []

	# --------------------------------------------------
	# Pre-group execution rows by Test ID
	# (fast lookup instead of repeated filtering)
	# --------------------------------------------------
	exec_by_test = df_exec.groupby("Test ID")

	for story_key, planned_tests in release_story_to_tests.items():

		# --------------------------------------------------
		# Unpack (release, story) safely
		# --------------------------------------------------
		if isinstance(story_key, (tuple, list)):
			release = normalise_text(story_key[0])
			story = normalise_id(story_key[1])
		else:
			story = normalise_id(story_key)

			release_val = story_to_release.get(story, "")
			if isinstance(release_val, (tuple, list)):
				release = normalise_text(release_val[0])
			else:
				release = normalise_text(release_val)

		# --------------------------------------------------
		# Iterate planned tests for this story
		# --------------------------------------------------
		for test_id in sorted(planned_tests):

			test_clean = normalise_id(test_id)

			# --------------------------------------------------
			# If test executed anywhere
			# --------------------------------------------------
			if test_clean in exec_by_test.groups:

				exec_rows = df_exec.loc[exec_by_test.groups[test_clean]]

				for _, r in exec_rows.iterrows():

					exec_story = normalise_id(r["Story"])

					# --------------------------------------------------
					# Alignment = executed under correct story?
					# --------------------------------------------------
					aligned = "YES" if exec_story == story else "NO"

					rows.append({
						"Release": release,
						"Story": story,
						"Test ID": test_clean,
						"Exec File": normalise_text(r["File"]),
						"Sheet": normalise_text(r["Sheet"]),
						# Prefer explicit Test Result if present
						"Status": r.get("Test Result", r.get("Status", "")),
						"Aligned": aligned,
						"Execution Location": exec_story,
					})

			else:
				# --------------------------------------------------
				# Test never executed anywhere
				# --------------------------------------------------
				rows.append({
					"Release": release,
					"Story": story,
					"Test ID": test_clean,
					"Exec File": "",
					"Sheet": "",
					"Status": "NOT EXECUTED",
					"Aligned": "NO",
					"Execution Location": "",
				})

	return pd.DataFrame(rows)

def _build_traceability_matrix(
	df_exec,
	release_story_to_tests,
	pass_values,
):
	"""
	Canonical flattened traceability dataset.

	One row per:
	Release -> Story -> Test ID

	Designed for:
	- filtering
	- audit
	- Power BI
	- machine-readable exports
	"""

	rows = []

	# --------------------------------------------------
	# Status classification
	# --------------------------------------------------
	df_exec = df_exec.copy()

	df_exec["Evidence"] = (
		df_exec["Evidence"]
		.astype(str)
		.str.lower()
	)

	df_exec["StatusClass"] = df_exec["Status"].apply(
		lambda s: classify_status(s, pass_values)
	)

	print("\nTRACEABILITY MATRIX DEBUG")

	print("release_story_to_tests sample:")
	for k in list(release_story_to_tests.keys())[:5]:
		print(k)

	print("\nExecution sample:")
	print(df_exec[["Story", "Test ID"]].head())

	# --------------------------------------------------
	# Fast lookup by Test ID
	# --------------------------------------------------
	exec_by_test = df_exec.groupby("Test ID")

	for story_key, planned_tests in sorted(release_story_to_tests.items()):

		release = normalise_text(story_key[0])
		story = normalise_id(story_key[1])

		for test_id in sorted(planned_tests):

			test_clean = normalise_id(test_id)

			# --------------------------------------------------
			# Executed somewhere?
			# --------------------------------------------------
			if test_clean in exec_by_test.groups:

				exec_rows = df_exec.loc[
					exec_by_test.groups[test_clean]
				]

				for _, r in exec_rows.iterrows():

					exec_story = normalise_id(r["Story"])

					aligned = exec_story == story

					traceability = (
						"ALIGNED"
						if aligned
						else "MISALIGNED"
					)

					rows.append({
						"Release": release,
						"Planned Story": story,
						"Test ID": test_clean,
						"Execution Story": exec_story,
						"Status": r["Status"],
						"Status Class": r["StatusClass"],
						"Evidence": r["Evidence"],
						"Aligned": "YES" if aligned else "NO",
						"Traceability Result": traceability,
						"Exec File": normalise_text(r["File"]),
						"Sheet": normalise_text(r["Sheet"]),
					})

			else:

				rows.append({
					"Release": release,
					"Planned Story": story,
					"Test ID": test_clean,
					"Execution Story": "",
					"Status": "NOT EXECUTED",
					"Status Class": "NOT_EXECUTED",
					"Evidence": "",
					"Aligned": "NO",
					"Traceability Result": "NOT_EXECUTED",
					"Exec File": "",
					"Sheet": "",
				})

	return pd.DataFrame(rows)

def _build_summary(release_story_to_tests, df_exec, pass_values, story_to_release):
	"""
	Build story-level summary.

	Assumptions:
	- df_exec already normalised by parser
	- IDs are clean (normalise_id applied upstream)
	- only light text normalisation needed here
	"""

	required_cols = {"Test ID", "Story", "Status", "Evidence"}
	missing = required_cols - set(df_exec.columns)
	if missing:
		raise ValueError(f"Missing required columns: {missing}")

	df_exec = df_exec.copy()

	# --------------------------------------------------
	# Evidence: normalise to lowercase for comparison only
	# (do NOT strip again – parser already handled whitespace)
	# --------------------------------------------------
	df_exec["Evidence"] = df_exec["Evidence"].astype(str).str.lower()

	# --------------------------------------------------
	# Status classification (single source of truth)
	# --------------------------------------------------
	df_exec["StatusClass"] = df_exec["Status"].apply(
		lambda s: classify_status(s, pass_values)
	)

	# --------------------------------------------------
	# GLOBAL execution set (IDs already normalised)
	# --------------------------------------------------
	all_executed = set(df_exec["Test ID"].dropna())

	# --------------------------------------------------
	# DUPLICATE DETECTION MAP (clean + no re-normalisation)
	# --------------------------------------------------
	test_to_stories = {}

	for test, story_val in zip(df_exec["Test ID"], df_exec["Story"]):
		if not test:
			continue

		# Only count real stories (ignore NEGATIVE, N/A, etc.)
		if isinstance(story_val, str) and story_val.startswith("STRY"):
			test_to_stories.setdefault(test, set()).add(story_val)

	duplicate_tests = {
		t: stories for t, stories in test_to_stories.items() if len(stories) > 1
	}

	summary_rows = []

	for story_key, planned_tests in sorted(release_story_to_tests.items()):

		# --------------------------------------------------
		# UNPACK (release, story)
		# --------------------------------------------------
		release = normalise_text(story_key[0])
		story = normalise_id(story_key[1])

		group = df_exec[df_exec["Story"] == story]

		planned = set(planned_tests)
		exec_tests = set(group["Test ID"].dropna())
		executed_in_story = exec_tests

		# --------------------------------------------------
		# TRACEABILITY
		# --------------------------------------------------
		genuine_missing = planned - all_executed
		misaligned = (planned - executed_in_story) & all_executed
		if genuine_missing:
			traceability = "🔴 Tests missing"
		elif misaligned:
			traceability = "🟡 Tests present (not linked)"
		else:
			traceability = "🟢 Tests present"

		# --------------------------------------------------
		# EXECUTION COUNT
		# --------------------------------------------------
		total_exec = len(executed_in_story)

		# --------------------------------------------------
		# STATUS COUNTS
		# --------------------------------------------------
		status_counts = group["StatusClass"].value_counts()

		passed = status_counts.get("PASS", 0)
		failed = status_counts.get("FAIL", 0)
		in_progress = status_counts.get("IN_PROGRESS", 0)
		not_started = status_counts.get("NOT_STARTED", 0)

		# --------------------------------------------------
		# EVIDENCE
		# --------------------------------------------------
		passed_with_evidence = len(group[
			(group["StatusClass"] == "PASS") &
			(group["Evidence"] == "yes")
		])

		# --------------------------------------------------
		# EXEC STATUS
		# --------------------------------------------------
		exec_status = _derive_exec_status(
			total_exec,
			passed,
			failed,
			in_progress,
			not_started,
			passed_with_evidence
		)

		# --------------------------------------------------
		# ISSUE BUILDING
		# --------------------------------------------------
		issues = []

		# Duplicate tests
		duplicates_for_story = [
			t for t, stories in duplicate_tests.items()
			if story in stories
		]

		if duplicates_for_story:
			dup_strings = []
			for t in sorted(duplicates_for_story):
				stories = sorted(s for s in duplicate_tests[t] if s != story)
				dup_strings.append(f"{t} (also in {', '.join(stories)})")

			issues.append(f"Duplicate: {'; '.join(dup_strings)}")

		# Traceability issues
		if genuine_missing:
			issues.append(f"Missing: {', '.join(sorted(genuine_missing))}")

		if misaligned:
			issues.append(f"Misaligned: {', '.join(sorted(misaligned))}")

		extra = executed_in_story - planned
		if extra:
			issues.append(f"Extra: {', '.join(sorted(extra))}")

		# Execution issues
		failed_tests = group[group["StatusClass"] == "FAIL"]["Test ID"].tolist()
		if failed_tests:
			issues.append(f"Failed: {', '.join(sorted(set(failed_tests)))}")

		missing_evidence_tests = group[
			(group["StatusClass"] == "PASS") &
			(group["Evidence"] == "no")
		]["Test ID"].tolist()

		if missing_evidence_tests:
			issues.append(f"Missing evidence: {', '.join(sorted(set(missing_evidence_tests)))}")

		issue_text = " | ".join(issues) if issues else ""

		# --------------------------------------------------
		# APPEND
		# --------------------------------------------------
		summary_rows.append({
			"Release": release,
			"Story": story,
			"Traceability": traceability,
			"Exec Status": exec_status,
			"Issue": issue_text,
			"Planned Tests": len(planned),
			"Execution Tests": total_exec,
			"Passed": int(passed),
			"Failed": int(failed),
			"In Progress": int(in_progress),
			"Not Started": int(not_started),
			"Passed w/ Evidence": int(passed_with_evidence)
		})
	
	# --------------------------------------------------
	# EXECUTION-ONLY STORIES
	# Stories present in execution but not in spec
	# --------------------------------------------------
	planned_stories = {
		normalise_id(story)
		for (_, story) in release_story_to_tests.keys()
	}

	executed_stories = {
		normalise_id(s)
		for s in df_exec["Story"].dropna()
		if isinstance(s, str) and s.startswith("STRY")
	}

	execution_only = executed_stories - planned_stories

	for story in sorted(execution_only):

		group = df_exec[df_exec["Story"] == story]

		executed_in_story = set(group["Test ID"].dropna())

		status_counts = group["StatusClass"].value_counts()

		passed = status_counts.get("PASS", 0)
		failed = status_counts.get("FAIL", 0)
		in_progress = status_counts.get("IN_PROGRESS", 0)
		not_started = status_counts.get("NOT_STARTED", 0)

		passed_with_evidence = len(group[
			(group["StatusClass"] == "PASS") &
			(group["Evidence"] == "yes")
		])

		exec_status = _derive_exec_status(
			len(executed_in_story),
			passed,
			failed,
			in_progress,
			not_started,
			passed_with_evidence
		)

		summary_rows.append({
			"Release": "UNREFERENCED",
			"Story": story,
			"Traceability": "🔴 Story missing from spec",
			"Exec Status": exec_status,
			"Issue": f"Execution-only story: {', '.join(sorted(executed_in_story))}",
			"Planned Tests": 0,
			"Execution Tests": len(executed_in_story),
			"Passed": int(passed),
			"Failed": int(failed),
			"In Progress": int(in_progress),
			"Not Started": int(not_started),
			"Passed w/ Evidence": int(passed_with_evidence)
		})

	return pd.DataFrame(summary_rows).sort_values(["Release", "Story"])

def _build_traceability_gaps(df_exec, release_story_to_tests, story_to_release):
	"""
	Build traceability gap analysis per story.

	Assumptions:
	- df_exec already normalised by parser
	- IDs are clean (normalise_id applied upstream)
	"""

	required_cols = {"Test ID", "Story", "Status", "Evidence"}
	missing = required_cols - set(df_exec.columns)

	if missing:
		raise ValueError(f"Missing required columns: {missing}")

	rows = []

	# --------------------------------------------------
	# GLOBAL execution set (already normalised IDs)
	# --------------------------------------------------
	all_exec_tests = set(df_exec["Test ID"].dropna())

	# --------------------------------------------------
	# STORY SETS
	# --------------------------------------------------
	planned_stories = {
		normalise_id(story)
		for (_, story) in release_story_to_tests.keys()
	}

	executed_stories = set(
		s for s in df_exec["Story"].dropna()
		if s and s.startswith("STRY")
	)

	# --------------------------------------------------
	# NORMAL planned story reconciliation
	# --------------------------------------------------
	for story_key, planned_tests in sorted(release_story_to_tests.items()):

		# --------------------------------------------------
		# UNPACK (release, story)
		# --------------------------------------------------
		release = normalise_text(story_key[0])
		story = normalise_id(story_key[1])

		group = df_exec[df_exec["Story"] == story]

		planned = set(planned_tests)
		exec_tests = set(group["Test ID"].dropna())

		# --------------------------------------------------
		# MISALIGNED (executed, but under different story)
		# --------------------------------------------------
		misaligned = (planned - exec_tests) & all_exec_tests

		# --------------------------------------------------
		# TRUE MISSING (not executed anywhere)
		# --------------------------------------------------
		genuine_missing = planned - all_exec_tests

		# --------------------------------------------------
		# STORY MISSING (includes misaligned)
		# --------------------------------------------------
		story_missing = genuine_missing.union(misaligned)

		# --------------------------------------------------
		# EXTRA (executed under this story but not planned)
		# --------------------------------------------------
		extra = exec_tests - planned

		# --------------------------------------------------
		# EXECUTION VIEW
		# --------------------------------------------------
		# What actually ran under this story
		execution_tests_display = exec_tests

		# What contributes to coverage (aligned + misaligned)
		execution_tests_coverage = (planned & exec_tests) | misaligned

		# --------------------------------------------------
		# SOURCE METADATA
		# --------------------------------------------------
		if not group.empty:
			source_file = normalise_text(group["File"].iloc[0])
			source_sheet = normalise_text(group["Sheet"].iloc[0])
		else:
			source_file = ""
			source_sheet = ""

		rows.append({
			"Release": release,
			"Story": story,
			"Source File": source_file,
			"Sheet": source_sheet,
			"Planned Tests": ", ".join(sorted(planned)),
			"Execution Tests": ", ".join(sorted(execution_tests_display)),
			"Missing Tests": ", ".join(sorted(story_missing)),
			"Misaligned Tests": ", ".join(sorted(misaligned)),
			"Extra Tests": ", ".join(sorted(extra)),
			"Planned Count": len(planned),
			"Coverage Count": len(execution_tests_coverage),
			"Missing Count": len(story_missing),
			"Extra Count": len(extra),
			"Has Gap": any([
				len(story_missing) > 0,
				len(extra) > 0,
				len(misaligned) > 0
			]),
		})

	# --------------------------------------------------
	# EXECUTION-ONLY STORIES
	# Stories present in execution but not in plan/spec
	# --------------------------------------------------
	execution_only = executed_stories - planned_stories

	for story in sorted(execution_only):

		group = df_exec[df_exec["Story"] == story]

		exec_tests = set(group["Test ID"].dropna())

		if not group.empty:
			source_file = normalise_text(group["File"].iloc[0])
			source_sheet = normalise_text(group["Sheet"].iloc[0])
		else:
			source_file = ""
			source_sheet = ""

		rows.append({
			"Release": "UNREFERENCED",
			"Story": story,
			"Source File": source_file,
			"Sheet": source_sheet,
			"Planned Tests": "",
			"Execution Tests": ", ".join(sorted(exec_tests)),
			"Missing Tests": "",
			"Misaligned Tests": "",
			"Extra Tests": ", ".join(sorted(exec_tests)),
			"Planned Count": 0,
			"Coverage Count": 0,
			"Missing Count": 0,
			"Extra Count": len(exec_tests),
			"Has Gap": True,
		})

	return pd.DataFrame(rows)

def write_output(
	output_path,
	plan_raw_rows,
	exec_rows,
	release_story_to_tests,
	result,
	story_to_release=None,
	df_exec=None,
	pass_values=None,
	include_audit=True,
	debug_dir=None
):
	"""
	Write reconciliation outputs to Excel (+ optional CSV debug dumps).

	Assumptions:
	- Parsers already normalise IDs/text
	- This layer should NOT re-clean IDs
	"""

	# --------------------------------------------------
	# Ensure df_exec exists (from parser or fallback)
	# --------------------------------------------------
	if df_exec is None:
		exec_rows_unique = list(dict.fromkeys(exec_rows))
		df_exec = pd.DataFrame(
			exec_rows_unique,
			columns=["Sheet", "Row", "Story", "Test ID", "Status", "File"],
		)
	else:
		df_exec = df_exec.copy()

	# --------------------------------------------------
	# Light column name normalisation ONLY (text-level)
	# --------------------------------------------------
	df_exec.columns = [normalise_text(c) for c in df_exec.columns]

	# --------------------------------------------------
	# Build outputs
	# --------------------------------------------------
	df_summary = _build_summary(
		release_story_to_tests,
		df_exec,
		pass_values,
		story_to_release,
	)

	df_gaps = _build_traceability_gaps(
		df_exec,
		release_story_to_tests,
		story_to_release,
	)

	df_matrix = _build_traceability_matrix(
		df_exec,
		release_story_to_tests,
		pass_values,
	)

	print("\nRELEASE STORY TEST DEBUG")

	for k, v in list(release_story_to_tests.items())[:5]:
		print(k)
		print(v)
		print(type(v))

	df_dashboard = _build_dashboard(
		df_summary,
		df_gaps,
		)

	df_missing = _df_from_set("MissingTest", result.missing_tests)
	df_extra = _df_from_set("ExtraTest", result.extra_tests)

	# --------------------------------------------------
	# Story → Test map
	# --------------------------------------------------
	st_rows = [
		(release, story, t)
		for (release, story), tests in sorted(release_story_to_tests.items())
		for t in sorted(tests)
	]

	df_story_map = pd.DataFrame(
		st_rows,
		columns=["Release", "Story", "Test"]
	)

	# --------------------------------------------------
	# Raw plan (no re-normalisation here)
	# --------------------------------------------------
	df_plan_raw = pd.DataFrame(
		plan_raw_rows,
		columns=["StoryCell", "RowText", "TestCell"],
	)

	# --------------------------------------------------
	# Ensure output path exists
	# --------------------------------------------------
	output_path = Path(output_path)
	output_path.parent.mkdir(parents=True, exist_ok=True)

	# --------------------------------------------------
	# Write Excel
	# --------------------------------------------------
	with pd.ExcelWriter(output_path, engine="openpyxl") as xw:

		df_dashboard.to_excel(
			xw,
			sheet_name="Dashboard",
			index=False,
		)

		df_summary.to_excel(
			xw,
			sheet_name="Summary",
			index=False,
		)

		df_matrix.to_excel(
			xw,
			sheet_name="Traceability_Matrix",
			index=False
		)
		
		print("\nMATRIX SHAPE:", df_matrix.shape)
		print(df_matrix.head(10))

		df_gaps.to_excel(xw, sheet_name="Traceability Gaps", index=False)
		df_missing.to_excel(xw, sheet_name="Missing", index=False)
		df_extra.to_excel(xw, sheet_name="Extra", index=False)
		df_story_map.to_excel(xw, sheet_name="Story_To_Test_Map", index=False)
		df_exec.to_excel(xw, sheet_name="Execution_Attachments", index=False)

		df_detail = _build_execution_detail(
			df_exec,
			release_story_to_tests,
			story_to_release,
		)
		df_detail.to_excel(xw, sheet_name="Execution_Detail", index=False)

		if include_audit:
			df_plan_raw.to_excel(xw, sheet_name="Plan_Raw", index=False)
			df_exec.to_excel(xw, sheet_name="Exec_Raw", index=False)

		# --------------------------------------------------
		# Workbook formatting
		# --------------------------------------------------
		_format_workbook(xw.book)

		# Hyperlinks
		_add_summary_hyperlinks(xw.book)

	# --------------------------------------------------
	# Optional debug outputs (CSV)
	# --------------------------------------------------
	if debug_dir:
		debug_dir = Path(debug_dir)
		debug_dir.mkdir(parents=True, exist_ok=True)

		df_summary.to_csv(debug_dir / "summary.csv", index=False)
		df_story_map.to_csv(debug_dir / "plan_extracted.csv", index=False)
		df_exec.to_csv(debug_dir / "exec_extracted.csv", index=False)
		df_missing.to_csv(debug_dir / "missing_tests.csv", index=False)
		df_extra.to_csv(debug_dir / "extra_tests.csv", index=False)


